# modules/operator/field_data_table.py

import streamlit as st
import pandas as pd
import io


from modules.analytics.repository import (
    get_analysis_by_system,
    delete_analysis
)


def render_field_data_table(df):

    # ======================================
    # TABLA COMPLETA
    # ======================================
    if df is not None and not df.empty:

        st.divider()

        title_col, btn_col = st.columns([4, 1])

        with title_col:

            st.subheader(
                "📋 Datos de campo"
            )

        with btn_col:

            # =========================
            # COPIA DF
            # =========================
            export_df = df.copy()

            # =========================
            # CLASIFICACIÓN
            # =========================
            def clasificar_lsi(lsi):

                if lsi > 0.5:
                    return "Incrustante"

                elif lsi < -0.5:
                    return "Corrosivo"

                else:
                    return "Ideal"

            
            export_df["Estado"] = export_df[
                "lsi_tablas"
            ].apply(
                clasificar_lsi
            )


            # ==================================
            # COLUMNAS EXPORTACIÓN
            # ==================================
            export_columns = [

                "sample_date",

                "ph",

                "tds_ppm",

                "temperature_c",

                "calcium_hardness",

                "alkalinity",

                "lsi_tablas",

                "Estado"

            ]

            # ==================================
            # COLUMNAS EXISTENTES
            # ==================================
            existing_columns = [

                col for col in export_columns

                if col in export_df.columns

            ]

            # ==================================
            # FILTRAR COLUMNAS
            # ==================================
            export_df = export_df[
                existing_columns
            ]

            # ==================================
            # VALIDAR DATOS
            # ==================================
            if export_df.empty:

                st.warning(
                    "No hay datos para exportar."
                )

                return

            # ==================================
            # EXCEL EN MEMORIA
            # ==================================
            output = io.BytesIO()

            with pd.ExcelWriter(

                output,

                engine="openpyxl"

            ) as writer:

                # ==============================
                # EXPORTAR EXCEL
                # ==============================
                export_df.to_excel(

                    writer,

                    index=False,

                    sheet_name="Historico"

                )

                workbook = writer.book

                worksheet = writer.sheets[
                    "Historico"
                ]

                # ==============================
                # COLORES
                # ==============================
                from openpyxl.styles import (
                    PatternFill
                )

                green_fill = PatternFill(

                    start_color="00FF00",

                    end_color="00FF00",

                    fill_type="solid"

                )

                red_fill = PatternFill(

                    start_color="FF0000",

                    end_color="FF0000",

                    fill_type="solid"

                )

                yellow_fill = PatternFill(

                    start_color="FFFF00",

                    end_color="FFFF00",

                    fill_type="solid"

                )

                # ==============================
                # ENCONTRAR COLUMNA ESTADO
                # ==============================
                estado_col = None

                for cell in worksheet[1]:

                    if cell.value == "Estado":

                        estado_col = cell.column

                # ==============================
                # COLOREAR CELDAS
                # ==============================
                if estado_col:

                    for row in range(
                        2,
                        worksheet.max_row + 1
                    ):

                        estado = worksheet.cell(

                            row=row,

                            column=estado_col

                        ).value

                        cell = worksheet.cell(

                            row=row,

                            column=estado_col

                        )

                        if estado == "Ideal":

                            cell.fill = green_fill

                        elif estado == "Incrustante":

                            cell.fill = red_fill

                        elif estado == "Corrosivo":

                            cell.fill = yellow_fill

            # ==================================
            # DOWNLOAD BUTTON
            # ==================================
            st.download_button(

                label="⬇️ Descargar",

                data=output.getvalue(),

                file_name="historico_lsi.xlsx",

                mime=(
                    "application/"
                    "vnd.openxmlformats-"
                    "officedocument."
                    "spreadsheetml.sheet"
                )

            )


        df_view = df.copy()
        # ======================================
        # RENOMBRAR COLUMNAS
        # ======================================
        df_view = df_view.rename(columns={

            "id": "ID",

            "sample_date": "Fecha",

            "ph": "pH",

            "temperature_c": "Temperature °C",

            "tds_ppm": "TDS ppm",

            "calcium_hardness": "Calcium Hardness",

            "alkalinity": "Alkalinity",

            "factor_a": "A_TDS",

            "factor_b": "B_Temp",

            "factor_c": "C_Hardness",

            "factor_d": "D_Alkalinity",

            "ph_s": "pH_s",

            "lsi": "LSI_log",

            "lsi_tablas": "LSI_tab"

        })

        # ======================================
        # ORDEN COLUMNAS
        # ======================================
        desired_cols = [
            "ID",
            "Fecha",
            "pH",
            "TDS ppm",
            "Temperature °C",
            "Calcium Hardness",
            "Alkalinity",
            "A_TDS",
            "B_Temp",
            "C_Hardness",
            "D_Alkalinity",
            "pH_s",
            "LSI_log",
            "LSI_tab",
        ]

        df_view = df_view[
            [c for c in desired_cols if c in df_view.columns]
        ]

        # ======================================
        # SOLO FECHA SIN HORA
        # ======================================
        df_view["Fecha"] = (
            df_view["Fecha"]
            .astype(str)
            .str[:10]
        )
        
        # ======================================
        # TABLA
        # ======================================
        st.dataframe(

            df_view,

            height=250
        )
                

        # ======================================
        # ELIMINAR REGISTRO
        # ======================================
        if not df_view.empty and "ID" in df_view.columns:

            st.divider()

            # ======================================
            # COLUMNAS
            # ======================================
            col1, col2, col3, col4 = st.columns(
                [1, 3, 1, 1]
            )

            # ======================================
            # IDS REALES SQLITE
            # ======================================
            record_ids = df_view["ID"].tolist()

            # ======================================
            # SELECTOR
            # ======================================
            with col1:

                #st.markdown(
                    #"### Selecciona un "
                #)
                selected_id = st.selectbox("",
                    options=record_ids,
                    key="selected_record_id",
                    format_func=lambda x: f"Registro #{x}"
                )

            # ======================================
            # ROW REAL
            # ======================================
            row = df_view[
                df_view["ID"] == selected_id
            ].iloc[0]

            # ======================================
            # INFO REGISTRO
            # ======================================
            with col2:

                st.markdown(
                    "### Selecciona registro para eliminar"
                )

                st.success(f"""

        |  Fecha: {str(row['Fecha']).split(' ')[0]}
         |  pH: {row['pH']}
         |  TDS: {row['TDS ppm']}
         |  Temp: {row['Temperature °C']}
         |  Ca: {row['Calcium Hardness']}
         |  Alk: {row['Alkalinity']}
        """)

            # ======================================
            # DELETE
            # ======================================
            with col3:

                st.markdown("### Acción")

                # ==============================
                # BOTÓN INICIAL
                # ==============================
                if st.button(
                    "🗑️ Eliminar",
                    key=f"delete_btn_{selected_id}"
                ):

                    st.session_state[
                        "confirm_delete"
                    ] = True

                # ==============================
                # CONFIRMACIÓN
                # ==============================
                if st.session_state.get(
                    "confirm_delete",
                    False
                ):

                    st.warning(
                        "¿Seguro que deseas eliminar este registro?"
                    )

                    col_yes, col_no = st.columns(2)

                    # ==========================
                    # CONFIRMAR
                    # ==========================
                    with col_yes:

                        if st.button(
                            "✅ Sí eliminar",
                            key=f"confirm_yes_{selected_id}"
                        ):

                            # ==========================
                            # DELETE DB
                            # ==========================
                            delete_analysis(
                                selected_id,
                                st.session_state.company_id
                            )

                            # ==========================
                            # RECARGAR DATOS DB
                            # ==========================
                            updated_df = get_analysis_by_system(

                                company_id=st.session_state.company_id,

                                system_id=st.session_state.system_id

                            )

                            # ==========================
                            # UPDATE SESSION DF
                            # ==========================
                            st.session_state[
                                "uploaded_df"
                            ] = updated_df

                            # ==========================
                            # SUCCESS
                            # ==========================
                            st.session_state[
                                "delete_success"
                            ] = True

                            # ==========================
                            # RESET CONFIRM
                            # ==========================
                            st.session_state[
                                "confirm_delete"
                            ] = False

                            # ==========================
                            # REFRESH UI
                            # ==========================
                            st.rerun()

                    # ==========================
                    # CANCELAR
                    # ==========================
                    with col_no:

                        if st.button(
                            "❌ Cancelar",
                            key=f"confirm_no_{selected_id}"
                        ):

                            st.session_state[
                                "confirm_delete"
                            ] = False

                            st.rerun()

            # ======================================
            # STATUS
            # ======================================
            with col4:

                st.markdown(
                    "### Estado"
                )

                if st.session_state.get(

                    "delete_success",

                    False
                ):

                    st.success(
                        "Registro eliminado"
                    )

                    st.session_state[
                        "delete_success"
                    ] = False