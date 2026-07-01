import io

import streamlit as st
import pandas as pd

from modules.analytics.repository import (
    get_analysis_by_system,
    delete_analysis
)

from modules.analytics.service import (
    clasificar_rsi
)


def render_rsi_field_data_table(df):

    if df is None or df.empty:

        st.info(
            "Sin registros guardados todavía"
        )

        return

    st.divider()

    title_col, btn_col = st.columns([4, 1])

    with title_col:

        st.subheader(
            "📋 Datos de campo"
        )

    with btn_col:

        export_df = df.copy()

        export_df["Estado"] = export_df[
            "rsi"
        ].apply(clasificar_rsi)

        export_columns = [
            "sample_date",
            "ph",
            "tds_ppm",
            "temperature_c",
            "calcium_hardness",
            "alkalinity",
            "rsi",
            "rsi_tablas",
            "Estado"
        ]

        existing_columns = [
            col for col in export_columns
            if col in export_df.columns
        ]

        export_df = export_df[existing_columns]

        if export_df.empty:

            st.warning("No hay datos para exportar.")

            return

        # ==================================
        # EXCEL EN MEMORIA
        # ==================================
        output = io.BytesIO()

        with pd.ExcelWriter(
            output,
            engine="openpyxl"
        ) as writer:

            export_df.to_excel(
                writer,
                index=False,
                sheet_name="Historico RSI"
            )

            worksheet = writer.sheets[
                "Historico RSI"
            ]

            from openpyxl.styles import PatternFill

            green_fill = PatternFill(
                start_color="00FF00",
                end_color="00FF00",
                fill_type="solid"
            )

            red_fill = PatternFill(
                start_color="FF0000",
                end_color="FF0000",
                fill_type="solid"
            )

            yellow_fill = PatternFill(
                start_color="FFFF00",
                end_color="FFFF00",
                fill_type="solid"
            )

            estado_col = None

            for cell in worksheet[1]:

                if cell.value == "Estado":

                    estado_col = cell.column

            if estado_col:

                for row in range(
                    2,
                    worksheet.max_row + 1
                ):

                    estado = worksheet.cell(
                        row=row,
                        column=estado_col
                    ).value

                    cell = worksheet.cell(
                        row=row,
                        column=estado_col
                    )

                    if estado == "Equilibrada":

                        cell.fill = green_fill

                    elif estado in (
                        "Incrustante",
                        "Muy incrustante"
                    ):

                        cell.fill = red_fill

                    elif estado in (
                        "Corrosiva",
                        "Muy corrosiva"
                    ):

                        cell.fill = yellow_fill

        st.download_button(

            label="⬇️ Descargar",

            data=output.getvalue(),

            file_name="historico_rsi.xlsx",

            mime=(
                "application/"
                "vnd.openxmlformats-"
                "officedocument."
                "spreadsheetml.sheet"
            )

        )

    # ======================================
    # TABLA
    # ======================================
    df_view = df.copy()

    df_view = df_view.rename(columns={

        "id": "ID",
        "sample_date": "Fecha",
        "ph": "pH",
        "temperature_c": "Temperature °C",
        "tds_ppm": "TDS ppm",
        "calcium_hardness": "Calcium Hardness",
        "alkalinity": "Alkalinity",
        "factor_a": "A_TDS",
        "factor_b": "B_Temp",
        "factor_c": "C_Hardness",
        "factor_d": "D_Alkalinity",
        "ph_s": "pH_s",
        "rsi": "RSI_log",
        "rsi_tablas": "RSI_tab"

    })

    desired_cols = [
        "ID",
        "Fecha",
        "pH",
        "TDS ppm",
        "Temperature °C",
        "Calcium Hardness",
        "Alkalinity",
        "A_TDS",
        "B_Temp",
        "C_Hardness",
        "D_Alkalinity",
        "pH_s",
        "RSI_log",
        "RSI_tab",
    ]

    df_view = df_view[
        [c for c in desired_cols if c in df_view.columns]
    ]

    if "Fecha" in df_view.columns:

        df_view["Fecha"] = (
            df_view["Fecha"]
            .astype(str)
            .str[:10]
        )

    st.dataframe(
        df_view,
        height=250
    )

    # ======================================
    # ELIMINAR REGISTRO
    # ======================================
    if not df_view.empty and "ID" in df_view.columns:

        st.divider()

        col1, col2, col3, col4 = st.columns(
            [1, 3, 1, 1]
        )

        record_ids = df_view["ID"].tolist()

        with col1:

            selected_id = st.selectbox(
                "",
                options=record_ids,
                key="rsi_selected_record_id",
                format_func=lambda x: f"Registro #{x}"
            )

        row = df_view[
            df_view["ID"] == selected_id
        ].iloc[0]

        with col2:

            st.markdown(
                "### Selecciona registro para eliminar"
            )

            st.success(f"""
    |  Fecha: {str(row.get('Fecha', '')).split(' ')[0]}
     |  pH: {row.get('pH', '')}
     |  TDS: {row.get('TDS ppm', '')}
     |  Temp: {row.get('Temperature °C', '')}
     |  Ca: {row.get('Calcium Hardness', '')}
     |  Alk: {row.get('Alkalinity', '')}
    """)

        with col3:

            st.markdown("### Acción")

            if st.button(
                "🗑️ Eliminar",
                key=f"rsi_delete_btn_{selected_id}"
            ):

                st.session_state[
                    "rsi_confirm_delete"
                ] = True

            if st.session_state.get(
                "rsi_confirm_delete",
                False
            ):

                st.warning(
                    "¿Seguro que deseas eliminar este registro?"
                )

                col_yes, col_no = st.columns(2)

                with col_yes:

                    if st.button(
                        "✅ Sí",
                        key=f"rsi_confirm_yes_{selected_id}"
                    ):

                        delete_analysis(
                            selected_id,
                            st.session_state.company_id
                        )

                        updated_df = get_analysis_by_system(
                            company_id=st.session_state.company_id,
                            system_id=st.session_state.system_id
                        )

                        st.session_state[
                            "uploaded_df"
                        ] = updated_df

                        st.session_state[
                            "rsi_delete_success"
                        ] = True

                        st.session_state[
                            "rsi_confirm_delete"
                        ] = False

                        st.rerun()

                with col_no:

                    if st.button(
                        "❌ Cancelar",
                        key=f"rsi_confirm_no_{selected_id}"
                    ):

                        st.session_state[
                            "rsi_confirm_delete"
                        ] = False

                        st.rerun()

        with col4:

            st.markdown("### Estado")

            if st.session_state.get(
                "rsi_delete_success",
                False
            ):

                st.success("Registro eliminado")

                st.session_state[
                    "rsi_delete_success"
                ] = False
